from fastapi import FastAPI, APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional
import uuid
from datetime import datetime, timezone
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")


# ============= MODELS =============

class AuctionConfig(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    num_teams: int = Field(default=8, ge=2, le=20)
    purse_amount: float = Field(default=100.0, gt=0)  # In Crores
    slots_per_team: int = Field(default=25, ge=1, le=50)
    is_configured: bool = Field(default=False)
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class AuctionConfigCreate(BaseModel):
    num_teams: int = Field(ge=2, le=20)
    purse_amount: float = Field(gt=0)
    slots_per_team: int = Field(ge=1, le=50)


class Team(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    logo_url: Optional[str] = None
    purse_remaining: float
    slots_filled: int = Field(default=0)
    total_slots: int
    total_spent: float = Field(default=0.0)
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class TeamCreate(BaseModel):
    name: str
    logo_url: Optional[str] = None


class TeamUpdate(BaseModel):
    name: Optional[str] = None
    logo_url: Optional[str] = None


class Player(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    role: str  # Batsman, Bowler, All-rounder, Wicketkeeper
    base_price: float  # In Lakhs
    status: str = Field(default="pending")  # pending, sold, unsold
    team_id: Optional[str] = None
    team_name: Optional[str] = None
    sale_price: Optional[float] = None  # In Lakhs
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class PlayerCreate(BaseModel):
    name: str
    role: str
    base_price: float


class PlayerSell(BaseModel):
    team_id: str
    sale_price: float


# ============= CONFIG ENDPOINTS =============

@api_router.get("/")
async def root():
    return {"message": "Cricket Auction API"}


@api_router.get("/config")
async def get_config():
    config = await db.auction_config.find_one({}, {"_id": 0})
    if not config:
        # Return default config
        return AuctionConfig().model_dump()
    if isinstance(config.get('created_at'), str):
        config['created_at'] = datetime.fromisoformat(config['created_at'])
    return config


@api_router.post("/config")
async def create_or_update_config(config_data: AuctionConfigCreate):
    existing = await db.auction_config.find_one({})
    
    if existing:
        # Update existing config
        await db.auction_config.update_one(
            {},
            {"$set": {
                "num_teams": config_data.num_teams,
                "purse_amount": config_data.purse_amount,
                "slots_per_team": config_data.slots_per_team,
                "is_configured": True
            }}
        )
        updated = await db.auction_config.find_one({}, {"_id": 0})
        return updated
    else:
        # Create new config
        config_obj = AuctionConfig(
            num_teams=config_data.num_teams,
            purse_amount=config_data.purse_amount,
            slots_per_team=config_data.slots_per_team,
            is_configured=True
        )
        doc = config_obj.model_dump()
        doc['created_at'] = doc['created_at'].isoformat()
        await db.auction_config.insert_one(doc)
        return config_obj.model_dump()


@api_router.delete("/config/reset")
async def reset_auction():
    """Reset entire auction - delete all data"""
    await db.auction_config.delete_many({})
    await db.teams.delete_many({})
    await db.players.delete_many({})
    return {"message": "Auction reset successfully"}


# ============= TEAM ENDPOINTS =============

@api_router.get("/teams", response_model=List[dict])
async def get_teams():
    teams = await db.teams.find({}, {"_id": 0}).to_list(100)
    for team in teams:
        if isinstance(team.get('created_at'), str):
            team['created_at'] = datetime.fromisoformat(team['created_at'])
    return teams


@api_router.post("/teams")
async def create_team(team_data: TeamCreate):
    # Check config
    config = await db.auction_config.find_one({})
    if not config or not config.get('is_configured'):
        raise HTTPException(status_code=400, detail="Please configure auction settings first")
    
    # Check team count
    team_count = await db.teams.count_documents({})
    if team_count >= config['num_teams']:
        raise HTTPException(status_code=400, detail=f"Maximum {config['num_teams']} teams allowed")
    
    # Check for duplicate name
    existing = await db.teams.find_one({"name": team_data.name})
    if existing:
        raise HTTPException(status_code=400, detail="Team name already exists")
    
    team_obj = Team(
        name=team_data.name,
        logo_url=team_data.logo_url,
        purse_remaining=config['purse_amount'],
        total_slots=config['slots_per_team']
    )
    
    doc = team_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.teams.insert_one(doc)
    
    return team_obj.model_dump()


@api_router.get("/teams/{team_id}")
async def get_team(team_id: str):
    team = await db.teams.find_one({"id": team_id}, {"_id": 0})
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    return team


@api_router.put("/teams/{team_id}")
async def update_team(team_id: str, team_data: TeamUpdate):
    team = await db.teams.find_one({"id": team_id})
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    
    update_data = {}
    if team_data.name is not None:
        # Check for duplicate name
        existing = await db.teams.find_one({"name": team_data.name, "id": {"$ne": team_id}})
        if existing:
            raise HTTPException(status_code=400, detail="Team name already exists")
        update_data['name'] = team_data.name
        # Update all players with this team
        await db.players.update_many(
            {"team_id": team_id},
            {"$set": {"team_name": team_data.name}}
        )
    
    if team_data.logo_url is not None:
        update_data['logo_url'] = team_data.logo_url
    
    if update_data:
        await db.teams.update_one({"id": team_id}, {"$set": update_data})
    
    updated = await db.teams.find_one({"id": team_id}, {"_id": 0})
    return updated


@api_router.delete("/teams/{team_id}")
async def delete_team(team_id: str):
    team = await db.teams.find_one({"id": team_id})
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    
    # Check if team has players
    player_count = await db.players.count_documents({"team_id": team_id})
    if player_count > 0:
        raise HTTPException(status_code=400, detail="Cannot delete team with players. Remove players first.")
    
    await db.teams.delete_one({"id": team_id})
    return {"message": "Team deleted successfully"}


# ============= PLAYER ENDPOINTS =============

@api_router.get("/players", response_model=List[dict])
async def get_players(role: Optional[str] = None, status: Optional[str] = None, team_id: Optional[str] = None):
    query = {}
    if role:
        query['role'] = role
    if status:
        query['status'] = status
    if team_id:
        query['team_id'] = team_id
    
    players = await db.players.find(query, {"_id": 0}).to_list(1000)
    for player in players:
        if isinstance(player.get('created_at'), str):
            player['created_at'] = datetime.fromisoformat(player['created_at'])
    return players


@api_router.post("/players")
async def create_player(player_data: PlayerCreate):
    if player_data.role not in ['Batsman', 'Bowler', 'All-rounder', 'Wicketkeeper']:
        raise HTTPException(status_code=400, detail="Invalid role")
    
    player_obj = Player(
        name=player_data.name,
        role=player_data.role,
        base_price=player_data.base_price
    )
    
    doc = player_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.players.insert_one(doc)
    
    return player_obj.model_dump()


@api_router.get("/players/{player_id}")
async def get_player(player_id: str):
    player = await db.players.find_one({"id": player_id}, {"_id": 0})
    if not player:
        raise HTTPException(status_code=404, detail="Player not found")
    return player


@api_router.delete("/players/{player_id}")
async def delete_player(player_id: str):
    player = await db.players.find_one({"id": player_id})
    if not player:
        raise HTTPException(status_code=404, detail="Player not found")
    
    # If player was sold, restore team's purse and slot
    if player['status'] == 'sold' and player.get('team_id'):
        await db.teams.update_one(
            {"id": player['team_id']},
            {
                "$inc": {
                    "purse_remaining": player['sale_price'],
                    "slots_filled": -1,
                    "total_spent": -player['sale_price']
                }
            }
        )
    
    await db.players.delete_one({"id": player_id})
    return {"message": "Player deleted successfully"}


@api_router.post("/players/{player_id}/sell")
async def sell_player(player_id: str, sell_data: PlayerSell):
    player = await db.players.find_one({"id": player_id})
    if not player:
        raise HTTPException(status_code=404, detail="Player not found")
    
    if player['status'] == 'sold':
        raise HTTPException(status_code=400, detail="Player already sold")
    
    team = await db.teams.find_one({"id": sell_data.team_id})
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    
    # Validate sale price
    if sell_data.sale_price < player['base_price']:
        raise HTTPException(status_code=400, detail="Sale price cannot be less than base price")
    
    # Convert sale price from Lakhs to Crores for comparison
    sale_price_crores = sell_data.sale_price / 100
    
    # Check purse
    if sale_price_crores > team['purse_remaining']:
        raise HTTPException(
            status_code=400, 
            detail=f"Insufficient purse. Team has {team['purse_remaining']:.2f} Cr remaining, but sale price is {sale_price_crores:.2f} Cr"
        )
    
    # Check slots
    if team['slots_filled'] >= team['total_slots']:
        raise HTTPException(
            status_code=400, 
            detail=f"No slots available. Team has filled all {team['total_slots']} slots"
        )
    
    # Update player
    await db.players.update_one(
        {"id": player_id},
        {"$set": {
            "status": "sold",
            "team_id": sell_data.team_id,
            "team_name": team['name'],
            "sale_price": sell_data.sale_price
        }}
    )
    
    # Update team
    await db.teams.update_one(
        {"id": sell_data.team_id},
        {
            "$inc": {
                "purse_remaining": -sale_price_crores,
                "slots_filled": 1,
                "total_spent": sale_price_crores
            }
        }
    )
    
    updated_player = await db.players.find_one({"id": player_id}, {"_id": 0})
    return updated_player


@api_router.post("/players/{player_id}/unsold")
async def mark_unsold(player_id: str):
    player = await db.players.find_one({"id": player_id})
    if not player:
        raise HTTPException(status_code=404, detail="Player not found")
    
    if player['status'] == 'sold':
        raise HTTPException(status_code=400, detail="Cannot mark sold player as unsold")
    
    await db.players.update_one(
        {"id": player_id},
        {"$set": {"status": "unsold"}}
    )
    
    updated_player = await db.players.find_one({"id": player_id}, {"_id": 0})
    return updated_player


@api_router.post("/players/{player_id}/reset")
async def reset_player(player_id: str):
    """Reset a player back to pending status"""
    player = await db.players.find_one({"id": player_id})
    if not player:
        raise HTTPException(status_code=404, detail="Player not found")
    
    # If player was sold, restore team's purse and slot
    if player['status'] == 'sold' and player.get('team_id'):
        sale_price_crores = player['sale_price'] / 100
        await db.teams.update_one(
            {"id": player['team_id']},
            {
                "$inc": {
                    "purse_remaining": sale_price_crores,
                    "slots_filled": -1,
                    "total_spent": -sale_price_crores
                }
            }
        )
    
    await db.players.update_one(
        {"id": player_id},
        {"$set": {
            "status": "pending",
            "team_id": None,
            "team_name": None,
            "sale_price": None
        }}
    )
    
    updated_player = await db.players.find_one({"id": player_id}, {"_id": 0})
    return updated_player


# ============= EXPORT ENDPOINT =============

@api_router.get("/export/excel")
async def export_excel():
    teams = await db.teams.find({}, {"_id": 0}).to_list(100)
    players = await db.players.find({"status": "sold"}, {"_id": 0}).to_list(1000)
    # Config available if needed for future enhancements
    _ = await db.auction_config.find_one({}, {"_id": 0})
    
    wb = Workbook()
    
    # Styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="007AFF", end_color="007AFF", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    summary_headers = ["Team Name", "Total Spent (Cr)", "Remaining Budget (Cr)", "Slots Filled", "Total Slots"]
    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    for row, team in enumerate(teams, 2):
        ws_summary.cell(row=row, column=1, value=team['name']).border = border
        ws_summary.cell(row=row, column=2, value=round(team.get('total_spent', 0), 2)).border = border
        ws_summary.cell(row=row, column=3, value=round(team.get('purse_remaining', 0), 2)).border = border
        ws_summary.cell(row=row, column=4, value=team.get('slots_filled', 0)).border = border
        ws_summary.cell(row=row, column=5, value=team.get('total_slots', 0)).border = border
    
    # Adjust column widths
    for col in range(1, 6):
        ws_summary.column_dimensions[chr(64 + col)].width = 20
    
    # Sheet 2: Master List
    ws_master = wb.create_sheet("All Players Sold")
    
    master_headers = ["Player Name", "Role", "Base Price (L)", "Sale Price (L)", "Team"]
    for col, header in enumerate(master_headers, 1):
        cell = ws_master.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    for row, player in enumerate(players, 2):
        ws_master.cell(row=row, column=1, value=player['name']).border = border
        ws_master.cell(row=row, column=2, value=player['role']).border = border
        ws_master.cell(row=row, column=3, value=player['base_price']).border = border
        ws_master.cell(row=row, column=4, value=player.get('sale_price', 0)).border = border
        ws_master.cell(row=row, column=5, value=player.get('team_name', '')).border = border
    
    for col in range(1, 6):
        ws_master.column_dimensions[chr(64 + col)].width = 20
    
    # Individual Team Sheets
    for team in teams:
        team_players = [p for p in players if p.get('team_id') == team['id']]
        
        # Sanitize sheet name (max 31 chars, no special chars)
        sheet_name = team['name'][:31].replace('/', '-').replace('\\', '-').replace('*', '-')
        ws_team = wb.create_sheet(sheet_name)
        
        team_headers = ["Player Name", "Role", "Base Price (L)", "Sale Price (L)"]
        for col, header in enumerate(team_headers, 1):
            cell = ws_team.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        for row, player in enumerate(team_players, 2):
            ws_team.cell(row=row, column=1, value=player['name']).border = border
            ws_team.cell(row=row, column=2, value=player['role']).border = border
            ws_team.cell(row=row, column=3, value=player['base_price']).border = border
            ws_team.cell(row=row, column=4, value=player.get('sale_price', 0)).border = border
        
        for col in range(1, 5):
            ws_team.column_dimensions[chr(64 + col)].width = 20
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=cricket_auction_export.xlsx"}
    )


# ============= DASHBOARD STATS =============

@api_router.get("/stats")
async def get_stats():
    config = await db.auction_config.find_one({}, {"_id": 0})
    teams = await db.teams.find({}, {"_id": 0}).to_list(100)
    
    total_players = await db.players.count_documents({})
    sold_players = await db.players.count_documents({"status": "sold"})
    unsold_players = await db.players.count_documents({"status": "unsold"})
    pending_players = await db.players.count_documents({"status": "pending"})
    
    # Role counts
    role_stats = {}
    for role in ['Batsman', 'Bowler', 'All-rounder', 'Wicketkeeper']:
        role_stats[role] = {
            "total": await db.players.count_documents({"role": role}),
            "sold": await db.players.count_documents({"role": role, "status": "sold"}),
            "unsold": await db.players.count_documents({"role": role, "status": "unsold"}),
            "pending": await db.players.count_documents({"role": role, "status": "pending"})
        }
    
    total_spent = sum(team.get('total_spent', 0) for team in teams)
    
    return {
        "config": config,
        "teams_count": len(teams),
        "total_players": total_players,
        "sold_players": sold_players,
        "unsold_players": unsold_players,
        "pending_players": pending_players,
        "role_stats": role_stats,
        "total_spent": round(total_spent, 2)
    }


# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
